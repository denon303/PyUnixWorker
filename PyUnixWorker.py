#==============================================================================
#title           :PyUnixWorker.py
#description     :Aplicacion consola para realizar tareas masivas mediante ssh en maquinas Unix/Linux.
#author          :Daniel Bejar Diaz
#date            :31/01/2019
#version         :0.1
#usage           :python PyUnixWorker.py
#notes           :
#python_version  :3.7.1  
#==============================================================================

import paramiko, time, os, sys
from paramiko_expect import SSHClientInteraction
from random import shuffle
import openpyxl

# PARAMIKO & NETVARS
global client
global remote_conn
global jumpNodesIndex
global state

# LOGIN VARIABLES
tUser = "t878xxx"
tPass = "xxxxxxxxx"
passwdPrompts = ["password:", "Password:", "password: ", "Password: " "password: ' ", "Password ' "]

# JUMPNODES DICTIONARIES...
jumpNode1 =  {  "hostname": "jump01", "ip": "10.0.0.1", "login" : "root" , "pass" : "madrid01", "jump": None, "strictoption" : 1, "Description" : "Salto Alternativa" } # Alternativa
jumpNode2 =   {  "hostname": "jump02",  "ip": "10.0.0.1" , "login" : tUser, "pass" : tPass, "jump": jump01, "Description" : "Salto Empresas" } #Salto Empresas
jumpNode3   =  {  "hostname": "jump03", "ip": "10.0.0.1", "login" : tUser , "pass" : tPass, "jump": None, "Description" : "Salto RIMA" } # Salto RIMA
jumpNode4 =  {  "hostname": "jump04", "ip": "10.0.0.1", "login" : tUser , "pass" : tPass, "jump": None, "Description" : "Salto NIMBA" } # Salto RIMA
jumpNode5 =  {  "hostname": "jump05", "ip": "10.0.0.1", "login" : tUser , "pass" : tPass, "jump": None, "Description" : "Salto NIMBA" } # Salto NIMBA
jumpNode6 =  {  "hostname": "jump06", "ip": "10.0.0.1", "login" : tUser , "pass" : tPass, "jump": jumpNode2, "Description" : "Salto Moviles" } # Salto Moviles 
jumpNode7 =  {  "hostname": "jump07", "ip": "10.0.0.1", "login" : tUser , "pass" : tPass, "jump": jumpNode2, "Description" : "Salto Moviles" } # Salto Moviles 

jumpNodes = [ jumpNode1, jumpNode2, jumpNode3, jumpNode4, jumpNode5, jumpNode6]

# TARGET SERVERS VARIABLES
targetServer = None

# EXCEL VARIABLES
inputExcl = openpyxl.load_workbook('C:/PETICION QUITAR EXPIRACION.xlsx') 
inputSheet= inputExcl["Hoja1"]
inventExcl = openpyxl.load_workbook('C:/InventarioUnix.xlsx', read_only = True, data_only = True)
inventSheet= inventExcl["t_inventario"]

shuffle(jumpNodes)
warnings.simplefilter("ignore")

class DevNull:
    def write(self, msg):
        pass

def StartWork():
    workcounter = 1
    for row in inputSheet.iter_rows(min_row=11, max_row=inputSheet.max_row, min_col=1, max_col=1, values_only=True):
        for cell in row:
            SearchInventory(workcounter, cell)
            time.sleep(1)
            workcounter += 1
            
def SearchInventory(number, hostname):
    print("\n *********************************************************************************************\n ")
    print(" *** Tarea:" + str(number) + " Hostname: " + hostname)
    print(" *** Buscando datos en el inventario para " + hostname + " ...")

    for row in inventSheet.iter_rows(min_row=1, max_row=inventSheet.max_row, min_col=1, max_col=1, values_only=False):
        for cell in row: 
            if str(hostname) == cell.value:
                targetServer = { "hostname": hostname, "ip": inventSheet.cell(row=cell.row, column=13).value, "pass": inventSheet.cell(row=cell.row, column=14).value }
                JumpNodeSearch(targetServer)
                break

def JumpNodeSearch(targetServer):
    global jumpNodesIndex
    global state
    state = "SEARCHJUMPNODE"
    print ("\n     Buscando Nodo de salto, puede tardar varios minutos...\n")
    tryJumpNode = 1

    for jumpNode in jumpNodes:
        if state == "LOGIN-PASSWD-OK":
            print("     Conectado con USER al servidor destino... State: " + state)
            break
        elif state == "LOGIN-PASSWD-FAIL":
            print("\n     *** Hay conectividad contra el servidor destino desde salto ------------ " + jumpNode["hostname"], jumpNode["ip"] + " pero la contraseña de: " + tUser + " es incorrecta... State: " + state)
            break
        elif state == "LOGIN-PASSWD-OK":
            print("\n     *** Hay conectividad contra el servidor destino desde salto ------------ " + jumpNode["hostname"], jumpNode["ip"] + " pero la contraseña de: " + tUser + " es incorrecta... State: " + state)
            break
        elif jumpNode["jump"] == None:
            ConnectToJumpNode(jumpNode, targetServer, tUser, tPass, tryJumpNode)
        else: # Falta por implementar el salto encadenado CHAIN-JUMP...
            pass
        time.sleep(1)
        tryJumpNode += 1
               
def GetOutputSSH(bufferSize, sleepTime):
    outputSplitLines = []
    time.sleep(sleepTime)
    output = remote_conn.recv(bufferSize)
    outputSplitLines = str(output).split("\\r\\n")  
    return outputSplitLines

def SendString(commandString, jumpNode, targetServer, hidePass):
    if hidePass:
        print ("      *** Enviando contraseña: " + "********"  + " Origen: " + jumpNode["hostname"] + " <-----> Destino: " + targetServer["hostname"])
    else:
        print ("      *** Ejecutando: " + commandString + " Origen: " + jumpNode["hostname"] + " <-----> Destino: " + targetServer["hostname"])
    remote_conn.send(commandString  + '\n')

def ConnectToJumpNode(jumpNode, targetServer, user, passwd, tryJumpNode):
    global state
    global client
    global remote_conn

    client = paramiko.SSHClient()
    client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    
    state = "CONNECT-TRY-JUMPNODE"
    
    try:
        print ("     " + str(tryJumpNode) + ") - Intentando autenticar con USER desde: " + jumpNode["hostname"] + " " + str(jumpNode["ip"]) + " Destino: " + str(targetServer["hostname"] + " " + str(targetServer["ip"])))
           
        client.connect(jumpNode["ip"], username=jumpNode["login"], password=jumpNode["pass"], timeout=8)
        remote_conn = client.invoke_shell()
        
        state = "CONNECTED-TO-JUMPNODE"
        
        if jumpNode["strictoption"] == 1:
            SendString('ssh -o StrictHostKeyChecking=no ' + user + '@' + str(targetServer["ip"]), jumpNode, targetServer, False)
        elif jumpNode["strictoption"] == 2:
            SendString('ssh ' + user + '@' + str(targetServer["ip"]), jumpNode, targetServer, False)
        
        if user != "root":
            LoginToServerWithTUSER(GetOutputSSH(5000,8), jumpNode, targetServer, user, passwd)
        else:
            pass
            #LoginToGetOutputSSH(1000,6), jumpNode, targetServer, user, passwd)
	    
    except paramiko.ssh_exception.AuthenticationException as e:
        print("### EXCEPTION Code(1)...")
        state = "CONNECT-FAIL-JUMPNODE"
    except paramiko.ssh_exception.BadHostKeyException as e:
        print("### EXCEPTION Code(2)...")
        state = "CONNECT-FAIL-JUMPNODE"
    except paramiko.ssh_exception.ChannelException as e:
        print("### EXCEPTION Code(3)...")
        state = "CONNECT-FAIL-JUMPNODE"
    except paramiko.ssh_exception.NoValidConnectionsError as e:
        print('SSH transport is not ready...')
        state = "CONNECT-FAIL-JUMPNODE"
    except paramiko.ssh_exception.PartialAuthentication as e:
        print("### EXCEPTION Code(5)...")
        state = "CONNECT-FAIL-JUMPNODE"
    except paramiko.ssh_exception.PasswordRequiredException as e:
        print("### EXCEPTION Code(6)...")
        state = "CONNECT-FAIL-JUMPNODE"
    except paramiko.ssh_exception.ProxyCommandFailure as e:
        print("### EXCEPTION Code(7)...")
        state = "CONNECT-FAIL-JUMPNODE"
    except paramiko.ssh_exception.SSHException as e:
        print('SSH transport is available!')
        state = "CONNECT-FAIL-JUMPNODE"
    finally:
        if client:
            client.close()

def LoginToServerWithTUSER(outputSplitLines, jumpNode, targetServer, user, passwd):
    global state
    state = "LOGIN-USER-PASSWD-PROMPT"
    #print(str(outputSplitLines[len(outputSplitLines)-1]))
    tempWords = str(outputSplitLines[len(outputSplitLines)-1]).split(" ")
    #print(tempWords)

    match = False
    
    for w in passwdPrompts:
        for h in tempWords:
            if h == w:
                match = True
                break
        else:
            pass
    
    if match == True:
        state = "TARGET-REACHED-FROM-JUMPNODE"
        AuthTUSER(jumpNode, targetServer, user, passwd)
        match = False
    else:
        state = "NO-RESPONSE-FROM-JUMPNODE"
           
def AuthTUSER(jumpNode, targetServer, user, passwd):
    global state
    state = "LOGIN-USER-PASSWD-SEND"
    SendString(passwd, jumpNode, targetServer, True)
    time.sleep(5)
    output = remote_conn.recv(5000)
    
    if str(output).find("DISPLAY=(") > 0:
       state = "LOGIN-USER-LOGPROMPT"
       SendString("\n\n", targetServer, False)
       time.sleep(2)
       output = remote_conn.recv(5000)
       if str(output).find("$") > 0: 
           print("\n     OK [TUSER] - Conectado como: " + user)
           state = "LOGIN-PASSWD-OK"
           #Worker()
       else:
           print("\n     NOK [TUSER] - Cambiando estado a: " + state)
           state = "LOGIN-PASSWD-FAIL"
           return 0
    else:
        if str(output).find("$") > 0: 
            print("     OK [TUSER] - Conectado como: " + user)
            state = "LOGIN-PASSWD-OK"
            #Worker()
        else:
            print("     NOK [TUSER] - Cambiando estado a: " + state)
            state = "LOGIN-PASSWD-FAIL"
            return 0
        
def Worker():
    """ PUT CODE FOR WORK ON EACH TARGETSERVER """
    pass

sys.stderr = DevNull()   
StartWork()
